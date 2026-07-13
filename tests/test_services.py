import json
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from case_builder_app.models import (
    INPUT_DATA_FAULT,
    INPUT_DATA_FREQUENCY_SWEEP,
    INPUT_DATA_SWITCHING,
    CasePart,
    CaseValue,
    ConditionalRule,
    ExclusionCombination,
    InputDataSettings,
    LayerChange,
    MMBlocks,
    MMLimit,
    ParameterChange,
    Project,
    ProjectSettings,
    RuleClause,
    SelectedCaseList,
)
from case_builder_app.services.export_service import ExportService
from case_builder_app.services.generator import GenerationService
from case_builder_app.services.input_data import input_data_cell_values
from case_builder_app.services.persistence import PROJECT_FILE_SUFFIX, PersistenceService


pytestmark = pytest.mark.filterwarnings("ignore:Data Validation extension is not supported.*:UserWarning")


def _conditional_rule_count(sheet) -> int:  # noqa: ANN001
    return sum(len(item.rules) for item in sheet.conditional_formatting)


def _conditional_rules(sheet):  # noqa: ANN001
    for item in sheet.conditional_formatting:
        for rule in item.rules:
            yield item.sqref, rule


def _conditional_fill_rgbs(sheet) -> set[str]:  # noqa: ANN001
    rgbs: set[str] = set()
    for _sqref, rule in _conditional_rules(sheet):
        dxf = rule.dxf
        if dxf is None or dxf.fill is None:
            continue
        for color in (dxf.fill.fgColor, dxf.fill.bgColor):
            if color.type == "rgb" and isinstance(color.rgb, str) and color.rgb != "00000000":
                rgbs.add(color.rgb)
    return rgbs


def _fill_key(cell) -> tuple:  # noqa: ANN001
    return (
        cell.fill.fill_type,
        _color_key(cell.fill.fgColor),
        _color_key(cell.fill.bgColor),
    )


def _color_key(color) -> tuple:  # noqa: ANN001
    if color.type == "rgb":
        value = color.rgb
    elif color.type == "theme":
        value = color.theme
    elif color.type == "indexed":
        value = color.indexed
    else:
        value = color.auto
    return (color.type, value, color.tint)


def _has_no_fill(cell) -> bool:  # noqa: ANN001
    return cell.fill.fill_type is None or _fill_key(cell) == (
        None,
        ("rgb", "00000000", 0.0),
        ("rgb", "00000000", 0.0),
    )


def _assert_input_data_static_page(
    workbook,
    export_path: Path,
    expected_top_values: list[str] | None = None,
) -> None:  # noqa: ANN001
    assert workbook["BaseSheet"].sheet_state == "hidden"
    assert [sheet.title for sheet in workbook.worksheets if sheet.sheet_state == "visible"][0] == "Input_Data"
    assert workbook.active.title == "Input_Data"

    input_data = workbook["Input_Data"]
    assert [input_data[f"B{row}"].value for row in range(2, 8)] == (
        expected_top_values or ["No", "Yes", "On", "Sequential", "No", "No"]
    )
    assert input_data.sheet_view.zoomScale == 86
    assert input_data.sheet_view.topLeftCell == "A40"
    assert input_data.column_dimensions["D"].width == 140
    assert input_data.row_dimensions[7].height == 14.65

    with ZipFile(export_path) as workbook_zip:
        input_data_xml = workbook_zip.read("xl/worksheets/sheet2.xml").decode("utf-8")
    assert 'x14:dataValidations count="6"' in input_data_xml
    assert "BaseSheet!$A$26:$A$31" in input_data_xml


def test_persistence_round_trip_uses_flat_exclusion_combinations(tmp_path: Path) -> None:
    part_s = CasePart(label="S")
    s1 = CaseValue(token="S1")
    s2 = CaseValue(token="S2")
    part_s.values = [s1, s2]

    part_v = CasePart(label="V")
    v1 = CaseValue(token="V1")
    v2 = CaseValue(token="V2")
    part_v.values = [v1, v2]

    rule = ConditionalRule(name="boost", priority=10)
    rule.clauses = [RuleClause(case_part_id=part_s.id, value_id=s2.id)]
    rule.changes.use_parameters = True
    rule.changes.parameters.append(ParameterChange(definition="Main", parameter="P_rule", value="9"))

    project = Project(
        name="Round Trip",
        settings=ProjectSettings(
            simple_export_enabled=True,
            input_data=InputDataSettings(studies=[INPUT_DATA_FAULT], frequency="50", mpe_workspace="Model.pswx"),
            case_name_order=[part_v.id, part_s.id],
        ),
        mm_blocks=MMBlocks(
            elements=["MM_230_OFT"],
            limits_by_voltage=[
                MMLimit(voltage="230", un="230", um="245", sdpf_lg="460", sdpf_ll="460", siwl_lg="850", siwl_ll="850", liwl="1050")
            ],
        ),
        case_parts=[part_s, part_v],
        conditional_rules=[rule],
        exclusions=[
            ExclusionCombination(
                clauses=[
                    RuleClause(case_part_id=part_s.id, value_id=s1.id),
                    RuleClause(case_part_id=part_v.id, value_id=v2.id),
                ]
            )
        ],
        selected_case_lists=[SelectedCaseList(name="Report", case_names=["S1_V1", "S2_V2"])],
    )

    service = PersistenceService()
    path = tmp_path / "round_trip"
    service.save_project(path, project)

    saved_path = tmp_path / f"round_trip{PROJECT_FILE_SUFFIX}"
    assert saved_path.exists()
    saved_payload = json.loads(saved_path.read_text(encoding="utf-8"))
    assert saved_payload["schema_version"] == 2
    assert "res_flux_enabled" not in saved_payload["settings"]
    assert saved_payload["settings"]["input_data"]["residual_flux"] == "No"

    loaded = service.load_project(saved_path)

    assert loaded.name == "Round Trip"
    assert loaded.settings.simple_export_enabled is True
    assert loaded.settings.input_data.normalized_studies() == [INPUT_DATA_FAULT]
    assert loaded.settings.input_data.frequency == "50"
    assert loaded.settings.input_data.mpe_workspace == "Model.pswx"
    assert loaded.settings.case_name_order == [part_v.id, part_s.id]
    assert loaded.mm_blocks.elements == ["MM_230_OFT"]
    assert loaded.mm_blocks.limits_by_voltage[0].voltage == "230"
    assert loaded.mm_blocks.limits_by_voltage[0].liwl == "1050"
    assert loaded.base_case.label == "Base Case"
    assert loaded.base_case.token == "BC"
    assert [part.label for part in loaded.case_parts] == ["S", "V"]
    assert len(loaded.exclusions) == 1
    assert len(loaded.selected_case_lists) == 1
    assert loaded.selected_case_lists[0].name == "Report"
    assert loaded.selected_case_lists[0].case_names == ["S1_V1", "S2_V2"]
    loaded_exclusion = loaded.exclusions[0]
    assert [(clause.case_part_id, clause.value_id) for clause in loaded_exclusion.clauses] == [
        (part_s.id, s1.id),
        (part_v.id, v2.id),
    ]
    assert len(loaded.conditional_rules) == 1
    assert loaded.conditional_rules[0].name == "boost"
    assert loaded.conditional_rules[0].changes.parameters[0].parameter == "P_rule"


def test_export_service_writes_expected_sheet_structure_and_order(tmp_path: Path) -> None:
    part_a = CasePart(label="A")
    a1 = CaseValue(token="A1")
    a1.changes.use_cb = True
    a1.changes.cb.on.append("CB_1")
    a1.changes.cb.off.append("CB_10")
    a1.changes.use_parameters = True
    a1.changes.parameters.extend(
        [
            ParameterChange(definition="Main", parameter="P_early", value="1"),
            ParameterChange(definition="Main", parameter="P_mid", value="2.5"),
            ParameterChange(definition="OnSS_VSR2", parameter="I_VSR", value="7"),
        ]
    )
    a1.changes.use_layers = True
    a1.changes.layers.extend(
        [
            LayerChange(section="Layers", layer_type="Main", target="L_MAIN_A", state="Enable"),
            LayerChange(section="Layers", layer_type="Extra", target="L_EXTRA_B", state="Disable"),
            LayerChange(section="Sweep_Components", layer_type="Main", target="SC_MAIN_B", state="Enable"),
            LayerChange(section="Sweep_Components", layer_type="Extra", target="SC_EXTRA_A", state="Disable"),
        ]
    )
    part_a.values = [a1]

    part_b = CasePart(label="B")
    b1 = CaseValue(token="B1")
    b1.changes.use_parameters = True
    b1.changes.parameters.extend(
        [
            ParameterChange(definition="Main", parameter="P_late", value="3"),
            ParameterChange(definition="OnSS_VSR1", parameter="I_VSR", value="6"),
        ]
    )
    part_b.values = [b1]

    rule = ConditionalRule(name="rule", priority=5)
    rule.clauses = [RuleClause(case_part_id=part_a.id, value_id=a1.id)]
    rule.changes.use_parameters = True
    rule.changes.parameters.append(ParameterChange(definition="Main", parameter="P_rule", value="4"))

    project = Project(
        settings=ProjectSettings(),
        case_parts=[part_a, part_b],
        conditional_rules=[rule],
    )
    project.base_case.changes.use_parameters = True
    project.base_case.changes.parameters.append(ParameterChange(definition="Main", parameter="P_base", value="0"))
    project.base_case.changes.fault_level.rpos = "0.1"
    project.base_case.changes.fault_level.xpos = "1"
    project.base_case.changes.fault_level.rzero = "2.5"
    project.base_case.changes.fault_level.xzero = "3"

    generated_cases, _stats = GenerationService().generate(project)
    export_path = tmp_path / "export.xlsx"
    ExportService().export(project, generated_cases, export_path)

    workbook = load_workbook(export_path)

    _assert_input_data_static_page(workbook, export_path)
    assert [sheet.title for sheet in workbook.worksheets if sheet.sheet_state == "visible"] == [
        "Input_Data",
        "Comp_Stat",
        "Layer_Stat",
        "Res_Flux",
        "MM_blocks",
    ]

    comp = workbook["Comp_Stat"]
    assert _conditional_rule_count(comp) >= 3
    assert {"FFA63A3A", "FFE46C0A", "FF77933C"} <= _conditional_fill_rgbs(comp)
    assert {"00A63A3A", "00E46C0A", "0077933C"}.isdisjoint(_conditional_fill_rgbs(comp))
    assert comp.freeze_panes == "C3"
    assert comp.sheet_view.zoomScale == 100
    assert comp["A1"].value == "Breaker_Status"
    assert all(_has_no_fill(comp.cell(1, column)) for column in range(1, 4))
    assert comp["A2"].value == "Definition"
    assert comp["B2"].value == "Breaker"
    assert comp["C2"].value == "A1_B1"
    assert comp["A3"].value == "CB_10"
    assert comp["B3"].value == "CB"
    assert comp["C3"].value == "OFF"
    assert comp["A4"].value == "CB_1"
    assert comp["B4"].value == "CB"
    assert comp["C4"].value == "ON"
    assert comp["A5"].value == "Fault_level"
    assert comp["A5"].font.color.rgb == "FFFF0000"
    assert comp["A5"].font.bold is True
    assert [(comp[f"A{row}"].value, comp[f"B{row}"].value) for row in range(6, 10)] == [
        ("Main", "Script_FL"),
        ("Main", "Script_FL"),
        ("Main", "Script_FL"),
        ("Main", "Script_FL"),
    ]
    assert [comp[f"C{row}"].value for row in range(6, 10)] == [0.1, 1, 2.5, 3]
    assert all(_has_no_fill(comp.cell(row, column)) for row in range(6, 10) for column in range(1, 4))
    assert comp["A10"].value == "Constants"
    assert comp["A10"].font.color.rgb == "FFFF0000"
    assert comp["A10"].font.bold is True
    assert [(comp[f"A{row}"].value, comp[f"B{row}"].value) for row in range(11, 18)] == [
        ("Main", "P_base"),
        ("Main", "P_early"),
        ("Main", "P_mid"),
        ("Main", "P_late"),
        ("Main", "P_rule"),
        ("OnSS_VSR1", "I_VSR"),
        ("OnSS_VSR2", "I_VSR"),
    ]
    assert [comp[f"C{row}"].value for row in range(11, 18)] == [0, 1, 2.5, 3, 4, 6, 7]

    layer = workbook["Layer_Stat"]
    assert _conditional_rule_count(layer) >= 2
    assert {"FFFFC7CE", "FFC6EFCE"} <= _conditional_fill_rgbs(layer)
    assert {"00FFC7CE", "00C6EFCE"}.isdisjoint(_conditional_fill_rgbs(layer))
    assert layer.freeze_panes == "C3"
    assert layer.sheet_view.zoomScale == 100
    assert layer.column_dimensions["A"].width == 15
    assert layer["A1"].value == "Layers"
    assert all(_has_no_fill(layer.cell(1, column)) for column in range(1, 4))
    assert layer["A2"].value == "Layer"
    assert layer["B2"].value == "String"
    assert [(layer[f"A{row}"].value, layer[f"B{row}"].value) for row in range(3, 5)] == [
        ("Extra", "L_EXTRA_B"),
        ("Main", "L_MAIN_A"),
    ]
    assert layer["A5"].value == "Sweep_Components"
    assert all(_has_no_fill(layer.cell(5, column)) for column in range(1, 4))
    assert [(layer[f"A{row}"].value, layer[f"B{row}"].value) for row in range(6, 8)] == [
        ("Extra", "SC_EXTRA_A"),
        ("Main", "SC_MAIN_B"),
    ]
    assert {tuple(rule.formula) for item in layer.conditional_formatting for rule in item.rules} >= {
        ('NOT(ISERROR(SEARCH("Disable",C3)))',),
        ('NOT(ISERROR(SEARCH("Enable",C3)))',),
        ('NOT(ISERROR(SEARCH("Disable",C6)))',),
        ('NOT(ISERROR(SEARCH("Enable",C6)))',),
    }
    state_rules = [rule for _sqref, rule in _conditional_rules(layer) if rule.text in {"Disable", "Enable"}]
    assert state_rules
    assert all(rule.dxf is not None and rule.dxf.font is None for rule in state_rules)


def test_export_service_writes_mm_blocks_by_voltage_order(tmp_path: Path) -> None:
    part = CasePart(label="S", values=[CaseValue(token="S1")])
    project = Project(
        settings=ProjectSettings(),
        case_parts=[part],
        mm_blocks=MMBlocks(
            elements=["MM_22_ONS", "MM_230_OFT", "MM_161_ONT"],
            limits_by_voltage=[
                MMLimit(voltage="22", un="22.8", um="24", sdpf_lg="50", sdpf_ll="50", siwl_lg="116", siwl_ll="116", liwl="145"),
                MMLimit(voltage="230", un="230", um="245", sdpf_lg="460", sdpf_ll="460", siwl_lg="850", siwl_ll="850", liwl="1050"),
                MMLimit(voltage="161", un="161", um="170", sdpf_lg="325", sdpf_ll="325", siwl_lg="620", siwl_ll="620", liwl="750"),
            ],
        ),
    )

    generated_cases, _stats = GenerationService().generate(project)
    export_path = tmp_path / "export_mm.xlsx"
    ExportService().export(project, generated_cases, export_path)

    workbook = load_workbook(export_path)

    _assert_input_data_static_page(workbook, export_path)
    assert [sheet.title for sheet in workbook.worksheets if sheet.sheet_state == "visible"] == [
        "Input_Data",
        "Comp_Stat",
        "Layer_Stat",
        "Res_Flux",
        "MM_blocks",
    ]
    sheet = workbook["MM_blocks"]
    assert [sheet.cell(1, column).value for column in range(1, 9)] == [
        "Group",
        "Un",
        "Um",
        "SDPF_LG",
        "SDPF_LL",
        "SIWL_LG",
        "SIWL_LL",
        "LIWL",
    ]
    assert [sheet[f"A{row}"].value for row in range(2, 5)] == ["MM_230_OFT", "MM_161_ONT", "MM_22_ONS"]
    assert [sheet[f"B{row}"].value for row in range(2, 5)] == [230, 161, 22.8]
    assert sheet["F4"].font.color.rgb in ("00000000", "FF000000")
    assert sheet["A2"].fill != sheet["A3"].fill


def test_export_service_preserves_template_styles_after_long_cb_section(tmp_path: Path) -> None:
    project = Project(
        settings=ProjectSettings(),
        case_parts=[CasePart(label="S", values=[CaseValue(token="S1")])],
    )
    project.base_case.changes.use_cb = True
    project.base_case.changes.cb.on = [f"CB_230_ITEM_{index:02d}" for index in range(1, 41)]
    project.base_case.changes.fault_level.rpos = "0.1"
    project.base_case.changes.fault_level.xpos = "1"
    project.base_case.changes.fault_level.rzero = "2"
    project.base_case.changes.fault_level.xzero = "3"
    project.base_case.changes.use_parameters = True
    project.base_case.changes.parameters.extend(
        [
            ParameterChange(definition="Main", parameter="AC_Vgridpu", value="1.05"),
            ParameterChange(definition="OnSS_VSR1", parameter="I_VSR", value="0.585"),
        ]
    )

    generated_cases, _stats = GenerationService().generate(project)
    export_path = tmp_path / "long_cb_export.xlsx"
    ExportService().export(project, generated_cases, export_path)

    workbook = load_workbook(export_path)
    template = load_workbook(Path("src/case_builder_app/assets/export_template.xlsx"))
    sheet = workbook["Comp_Stat"]
    template_sheet = template["Comp_Stat"]

    fault_row = next(row for row in range(1, sheet.max_row + 1) if sheet[f"A{row}"].value == "Fault_level")
    constants_row = next(row for row in range(1, sheet.max_row + 1) if sheet[f"A{row}"].value == "Constants")

    assert fault_row > 33
    assert sheet.cell(fault_row, 1).font.color.rgb == "FFFF0000"
    assert sheet.cell(fault_row, 1).font.bold is True
    assert _fill_key(sheet.cell(fault_row, 1)) == _fill_key(template_sheet["A33"])
    assert sheet.cell(constants_row, 1).font.color.rgb == "FFFF0000"
    assert sheet.cell(constants_row, 1).font.bold is True
    assert _fill_key(sheet.cell(constants_row, 1)) == _fill_key(template_sheet["A38"])

    for row in range(fault_row + 1, fault_row + 5):
        assert (sheet[f"A{row}"].value, sheet[f"B{row}"].value) == ("Main", "Script_FL")
        assert sheet[f"A{row}"].font.bold is False
        assert sheet[f"B{row}"].font.bold is False
        assert all(_has_no_fill(sheet.cell(row, column)) for column in range(1, 4))

    main_constant_row = constants_row + 1
    other_constant_row = constants_row + 2
    assert (sheet[f"A{main_constant_row}"].value, sheet[f"B{main_constant_row}"].value) == ("Main", "AC_Vgridpu")
    assert (sheet[f"A{other_constant_row}"].value, sheet[f"B{other_constant_row}"].value) == ("OnSS_VSR1", "I_VSR")
    assert _fill_key(sheet[f"B{main_constant_row}"]) == _fill_key(template_sheet["B39"])
    assert _fill_key(sheet[f"C{main_constant_row}"]) == _fill_key(template_sheet["C39"])
    assert _fill_key(sheet[f"B{other_constant_row}"]) == _fill_key(template_sheet["B40"])
    assert _fill_key(sheet[f"C{other_constant_row}"]) == _fill_key(template_sheet["C40"])


def test_export_service_real_energisation_project_keeps_excel_styles(tmp_path: Path) -> None:
    project_path = Path("F4_Energisation_FS.casebuilder.json")
    if not project_path.exists():
        pytest.skip("F4_Energisation_FS.casebuilder.json fixture is not present in this workspace")
    project = PersistenceService().load_project(project_path)
    project.settings.input_data = InputDataSettings(studies=[INPUT_DATA_SWITCHING], residual_flux="Yes")
    project.base_case.changes.use_flux = True
    generated_cases, _stats = GenerationService().generate(project)
    export_path = tmp_path / "energisation_export.xlsx"
    ExportService().export(project, generated_cases, export_path)

    workbook = load_workbook(export_path)

    _assert_input_data_static_page(workbook, export_path, ["No", "No", "On", "Sequential", "No", "Yes"])
    assert [sheet.title for sheet in workbook.worksheets if sheet.sheet_view.tabSelected] == ["Input_Data"]

    comp = workbook["Comp_Stat"]
    cb_rules = {
        rule.text: rule
        for sqref, rule in _conditional_rules(comp)
        if str(sqref).startswith("C3:") and rule.type == "containsText" and rule.text in {"OFF", "SWITCH", "ON"}
    }
    assert set(cb_rules) == {"OFF", "SWITCH", "ON"}
    assert cb_rules["OFF"].dxf.fill.bgColor.rgb == "FFA63A3A"
    assert cb_rules["SWITCH"].dxf.fill.bgColor.rgb == "FFE46C0A"
    assert cb_rules["ON"].dxf.fill.bgColor.rgb == "FF77933C"
    assert all(rule.dxf.fill.fgColor.rgb == "00000000" for rule in cb_rules.values())
    assert all(rule.dxf.font.color.type == "theme" and rule.dxf.font.color.theme == 0 for rule in cb_rules.values())

    layer = workbook["Layer_Stat"]
    layer_rules = {
        rule.text: rule
        for sqref, rule in _conditional_rules(layer)
        if str(sqref).startswith("C3:") and rule.type == "containsText"
    }
    sweep_rules = {
        rule.text: rule
        for sqref, rule in _conditional_rules(layer)
        if str(sqref).startswith("C13:") and rule.type == "containsText"
    }
    assert layer_rules["Disable"].dxf.fill.bgColor.rgb == "FFFFC7CE"
    assert layer_rules["Enable"].dxf.fill.bgColor.rgb == "FFC6EFCE"
    assert sweep_rules["Disable"].dxf.fill.bgColor.theme == 9
    assert sweep_rules["Disable"].dxf.fill.bgColor.tint == 0.3999450666829432
    assert sweep_rules["Enable"].dxf.fill.bgColor.theme == 5
    assert sweep_rules["Enable"].dxf.fill.bgColor.tint == 0.0
    assert all(rule.dxf.font is None for rule in [*layer_rules.values(), *sweep_rules.values()])

    flux = workbook["Res_Flux"]
    assert _conditional_rule_count(flux) == 1
    assert flux["C3"].font.name == "Calibri"
    assert flux["C3"].font.sz == 9
    assert flux["C3"].font.bold is False
    assert (flux["C3"].border.left.style, flux["C3"].border.right.style, flux["C3"].border.top.style, flux["C3"].border.bottom.style) == (
        "medium",
        "thin",
        "medium",
        "thin",
    )
    assert (flux["D3"].border.left.style, flux["D3"].border.right.style, flux["D3"].border.top.style, flux["D3"].border.bottom.style) == (
        "thin",
        "medium",
        "medium",
        "thin",
    )
    assert (flux["C4"].border.left.style, flux["C4"].border.right.style, flux["C4"].border.top.style, flux["C4"].border.bottom.style) == (
        "medium",
        "thin",
        "thin",
        "medium",
    )
    assert (flux["D4"].border.left.style, flux["D4"].border.right.style, flux["D4"].border.top.style, flux["D4"].border.bottom.style) == (
        "thin",
        "medium",
        "thin",
        "medium",
    )


def test_export_service_simple_export_writes_plain_workbook(tmp_path: Path) -> None:
    project = Project(
        settings=ProjectSettings(simple_export_enabled=True),
        case_parts=[CasePart(label="S", values=[CaseValue(token="S1")])],
    )
    project.base_case.changes.use_cb = True
    project.base_case.changes.cb.on = ["CB_230_A"]

    generated_cases, _stats = GenerationService().generate(project)
    export_path = tmp_path / "simple_export.xlsx"
    ExportService().export(project, generated_cases, export_path)

    workbook = load_workbook(export_path)

    _assert_input_data_static_page(workbook, export_path)
    assert workbook.sheetnames == ["BaseSheet", "Input_Data", "Comp_Stat", "Layer_Stat", "Res_Flux", "MM_blocks"]
    assert workbook["Comp_Stat"]["A1"].value == "Breaker_Status"
    assert workbook["Comp_Stat"]["A3"].value == "CB_230_A"
    assert len(workbook["Comp_Stat"].conditional_formatting) == 0


def test_export_service_applies_project_input_data_settings(tmp_path: Path) -> None:
    project = Project(
        settings=ProjectSettings(
            input_data=InputDataSettings(
                studies=[INPUT_DATA_SWITCHING, INPUT_DATA_FAULT],
                frequency="50",
                initialisation_duration="3.5",
                time_step="5",
                plot_step="50",
                create_cases="No",
                take_snapshot="No",
                adjust_names="No",
                snapshot_time="3",
                mpe_workspace="MPE_Model.pswx",
                pscad_version="5.0.1",
                fortran_compiler="GFortran 4.6.2",
                parallel_simulations="6",
                pscad_instances="7",
                final_duration="0.5",
                switch_operation="On-Off",
                switch_type="Sequential",
                points_over_wave="25",
                switch_start="2.2",
                switch_points_to_check="11",
                second_switch_delay="0.2",
                fault_start="2.1",
                fault_points_to_check="9",
            ),
        ),
        case_parts=[CasePart(label="S", values=[CaseValue(token="S1")])],
    )

    generated_cases, _stats = GenerationService().generate(project)
    export_path = tmp_path / "custom_input_data.xlsx"
    ExportService().export(project, generated_cases, export_path)

    input_data = load_workbook(export_path, data_only=False)["Input_Data"]
    input_data_values = load_workbook(export_path, data_only=True)["Input_Data"]

    assert input_data["B2"].value == "No"
    assert input_data["B3"].value == "No"
    assert input_data["B4"].value == "On-Off"
    assert input_data["B5"].value == "Sequential"
    assert input_data["B6"].value == "Yes"
    assert input_data["B7"].value == "No"
    assert input_data["B16"].value == 50
    assert input_data["B17"].value == 3.5
    assert input_data["B18"].value == 5
    assert input_data["B19"].value == 50
    assert input_data["B20"].value == "No"
    assert input_data["B21"].value == "No"
    assert input_data["B22"].value == "No"
    assert input_data["B23"].value == 3
    assert input_data["B24"].value == "No"
    assert input_data["B25"].value == "MPE_Model.pswx"
    assert input_data["B26"].value is None
    assert input_data["B27"].value == "5.0.1"
    assert input_data["B28"].value == "GFortran 4.6.2"
    assert input_data["B29"].value == 6
    assert input_data["B30"].value == 7
    assert input_data["B36"].value == 0.5
    assert input_data["B39"].value == 2.1
    assert input_data["B40"].value == "=1/(B16*25)"
    assert input_data["B41"].value == "=B39+11*B40"
    assert input_data_values["B40"].value == pytest.approx(0.0008)
    assert input_data_values["B41"].value == pytest.approx(2.1088)
    assert input_data["B44"].value == 2.2
    assert input_data["B45"].value == "=1/(B16*25)"
    assert input_data["B46"].value == "=B44+11*B45"
    assert input_data_values["B45"].value == pytest.approx(0.0008)
    assert input_data_values["B46"].value == pytest.approx(2.2088)
    assert input_data["B47"].value == 0.2


def test_input_data_residual_flux_is_only_enabled_for_switching_studies() -> None:
    switching_project = Project(
        settings=ProjectSettings(
            input_data=InputDataSettings(studies=[INPUT_DATA_SWITCHING], residual_flux="Yes"),
        )
    )
    fault_project = Project(
        settings=ProjectSettings(
            input_data=InputDataSettings(studies=[INPUT_DATA_FAULT], residual_flux="Yes"),
        )
    )
    sweep_project = Project(
        settings=ProjectSettings(
            input_data=InputDataSettings(studies=[INPUT_DATA_FREQUENCY_SWEEP], residual_flux="Yes"),
        )
    )

    assert input_data_cell_values(switching_project)["B7"] == "Yes"
    assert input_data_cell_values(fault_project)["B7"] == "No"
    assert input_data_cell_values(sweep_project)["B7"] == "No"


def test_export_service_sorts_cb_rows_by_voltage_then_natural_name(tmp_path: Path) -> None:
    project = Project(
        settings=ProjectSettings(),
        case_parts=[CasePart(label="S", values=[CaseValue(token="S1")])],
    )
    project.base_case.changes.use_cb = True
    project.base_case.changes.cb.on = [
        "CB_66_A",
        "CB_230_OFT10",
        "CB_230_HF2",
        "CB_230_HF1",
        "CB_230_OFT2",
    ]

    generated_cases, _stats = GenerationService().generate(project)
    export_path = tmp_path / "export_cb_order.xlsx"
    ExportService().export(project, generated_cases, export_path)

    sheet = load_workbook(export_path)["Comp_Stat"]

    assert [sheet[f"A{row}"].value for row in range(3, 8)] == [
        "CB_230_HF1",
        "CB_230_HF2",
        "CB_230_OFT2",
        "CB_230_OFT10",
        "CB_66_A",
    ]


def test_export_service_uses_distinct_cb_group_fills_beyond_template_palette(tmp_path: Path) -> None:
    project = Project(
        settings=ProjectSettings(),
        case_parts=[CasePart(label="S", values=[CaseValue(token="S1")])],
    )
    project.base_case.changes.use_cb = True
    project.base_case.changes.cb.on = [
        "CB_500_A",
        "CB_400_A",
        "CB_330_A",
        "CB_230_A",
        "CB_161_A",
        "CB_110_A",
        "CB_66_A",
    ]

    generated_cases, _stats = GenerationService().generate(project)
    export_path = tmp_path / "export_many_voltage_groups.xlsx"
    ExportService().export(project, generated_cases, export_path)

    sheet = load_workbook(export_path)["Comp_Stat"]
    fills = [_fill_key(sheet[f"A{row}"]) for row in range(3, 10)]

    assert [sheet[f"A{row}"].value for row in range(3, 10)] == [
        "CB_500_A",
        "CB_400_A",
        "CB_330_A",
        "CB_230_A",
        "CB_161_A",
        "CB_110_A",
        "CB_66_A",
    ]
    assert len(set(fills)) == len(fills)
    assert fills[0] != fills[5]


def test_export_service_uses_same_voltage_fills_for_cb_and_mm_blocks(tmp_path: Path) -> None:
    project = Project(
        settings=ProjectSettings(),
        case_parts=[CasePart(label="S", values=[CaseValue(token="S1")])],
        mm_blocks=MMBlocks(
            elements=[
                "MM_500_A",
                "MM_400_A",
                "MM_330_A",
                "MM_230_A",
                "MM_161_A",
                "MM_110_A",
                "MM_66_A",
            ],
            limits_by_voltage=[
                MMLimit(voltage=voltage, un=voltage)
                for voltage in ("500", "400", "330", "230", "161", "110", "66")
            ],
        ),
    )
    project.base_case.changes.use_cb = True
    project.base_case.changes.cb.on = [
        "CB_500_A",
        "CB_400_A",
        "CB_330_A",
        "CB_230_A",
        "CB_161_A",
        "CB_110_A",
        "CB_66_A",
    ]

    generated_cases, _stats = GenerationService().generate(project)
    export_path = tmp_path / "export_shared_voltage_fills.xlsx"
    ExportService().export(project, generated_cases, export_path)

    workbook = load_workbook(export_path)
    comp = workbook["Comp_Stat"]
    mm = workbook["MM_blocks"]
    cb_fills = {comp[f"A{row}"].value.split("_")[1]: _fill_key(comp[f"A{row}"]) for row in range(3, 10)}
    mm_fills = {mm[f"A{row}"].value.split("_")[1]: _fill_key(mm[f"A{row}"]) for row in range(2, 9)}

    assert cb_fills == mm_fills
    assert len(set(mm_fills.values())) == len(mm_fills)
