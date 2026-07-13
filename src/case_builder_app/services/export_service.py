from __future__ import annotations

from importlib import resources
from pathlib import Path
import warnings
from zipfile import ZIP_DEFLATED, ZipFile
import os
import re
import tempfile
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from case_builder_app.models import Project
from case_builder_app.services.export_ordering import (
    ordered_breaker_keys,
    ordered_flux_keys,
    ordered_layer_keys,
    ordered_parameter_keys,
)
from case_builder_app.services.generator import GeneratedCase
from case_builder_app.services.input_data import apply_input_data_settings, input_data_cached_formula_values
from case_builder_app.services.sheet_builders import (
    ExportRow,
    build_comp_stat_rows,
    build_layer_stat_rows,
    build_mm_blocks_rows,
    build_res_flux_rows,
    write_simple_rows,
)
from case_builder_app.services.style_template import (
    StyleSheet,
    add_cb_conditional_formatting,
    add_constants_conditional_formatting,
    add_duplicate_case_name_formatting,
    add_layer_conditional_formatting,
    build_voltage_fill_map,
    prepare_sheet,
    select_single_sheet,
    set_common_widths,
    set_workbook_zoom,
    trim_sheet,
    write_body_row,
    write_header_row,
    write_section_row,
)


TEMPLATE_NAME = "export_template.xlsx"
INPUT_DATA_SHEET_NAME = "Input_Data"
WORKBOOK_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
WORKSHEET_REL_TYPE = f"{OFFICE_REL_NS}/worksheet"
XR_NS = "http://schemas.microsoft.com/office/spreadsheetml/2014/revision"
EXT_LST_RE = re.compile(rb"<extLst>.*?</extLst>", re.DOTALL)


class ExportService:
    def export(self, project: Project, generated_cases: list[GeneratedCase], path: str | Path) -> None:
        if project.settings.simple_export_enabled:
            self._export_simple(project, generated_cases, path)
            return
        self._export_styled(project, generated_cases, path)

    def _export_styled(self, project: Project, generated_cases: list[GeneratedCase], path: str | Path) -> None:
        case_names = [item.name for item in generated_cases]
        breaker_keys = ordered_breaker_keys(project, generated_cases)
        parameter_keys = ordered_parameter_keys(project, generated_cases)
        layer_keys = ordered_layer_keys(project, generated_cases)
        flux_keys = ordered_flux_keys(project, generated_cases)
        include_fault = any(item.fault_level for item in generated_cases)

        workbook, input_data_extension = self._load_template_workbook()
        if INPUT_DATA_SHEET_NAME in workbook.sheetnames:
            apply_input_data_settings(workbook[INPUT_DATA_SHEET_NAME], project)

        voltage_fills = build_voltage_fill_map(
            workbook["Comp_Stat"],
            [*breaker_keys, *project.mm_blocks.elements],
        )

        comp_rows = build_comp_stat_rows(case_names, generated_cases, breaker_keys, include_fault, parameter_keys)
        self._build_comp_stat_sheet(workbook["Comp_Stat"], comp_rows, len(case_names), voltage_fills)

        layer_rows = build_layer_stat_rows(case_names, generated_cases, layer_keys)
        self._build_layer_stat_sheet(workbook["Layer_Stat"], layer_rows, len(case_names))

        if "Res_Flux" in workbook.sheetnames:
            flux_rows = build_res_flux_rows(case_names, generated_cases, flux_keys)
            self._build_res_flux_sheet(workbook["Res_Flux"], flux_rows, len(case_names), workbook["Comp_Stat"])

        if "MM_blocks" in workbook.sheetnames:
            self._build_mm_blocks_sheet(workbook["MM_blocks"], build_mm_blocks_rows(project), voltage_fills)

        self._save_export(workbook, project, path, input_data_extension)

    def _export_simple(self, project: Project, generated_cases: list[GeneratedCase], path: str | Path) -> None:
        workbook, input_data_extension = self._load_template_workbook()
        if INPUT_DATA_SHEET_NAME in workbook.sheetnames:
            apply_input_data_settings(workbook[INPUT_DATA_SHEET_NAME], project)
        for sheet_name in ("Comp_Stat", "Layer_Stat", "Res_Flux", "MM_blocks"):
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]

        case_names = [item.name for item in generated_cases]
        breaker_keys = ordered_breaker_keys(project, generated_cases)
        parameter_keys = ordered_parameter_keys(project, generated_cases)
        layer_keys = ordered_layer_keys(project, generated_cases)
        flux_keys = ordered_flux_keys(project, generated_cases)
        include_fault = any(item.fault_level for item in generated_cases)

        write_simple_rows(
            workbook.create_sheet("Comp_Stat"),
            build_comp_stat_rows(case_names, generated_cases, breaker_keys, include_fault, parameter_keys),
        )
        write_simple_rows(
            workbook.create_sheet("Layer_Stat"),
            build_layer_stat_rows(case_names, generated_cases, layer_keys),
        )
        write_simple_rows(
            workbook.create_sheet("Res_Flux"),
            build_res_flux_rows(case_names, generated_cases, flux_keys),
        )
        write_simple_rows(workbook.create_sheet("MM_blocks"), build_mm_blocks_rows(project))

        self._save_export(workbook, project, path, input_data_extension)

    def _load_template_workbook(self) -> tuple[Workbook, bytes | None]:
        with resources.as_file(resources.files("case_builder_app.assets").joinpath(TEMPLATE_NAME)) as template_path:
            input_data_extension = _extract_input_data_extension(template_path)
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", message="Data Validation extension is not supported.*")
                return load_workbook(template_path), input_data_extension

    def _save_export(
        self,
        workbook: Workbook,
        project: Project,
        path: str | Path,
        input_data_extension: bytes | None,
    ) -> None:
        self._select_export_active_sheet(workbook)
        set_workbook_zoom(
            workbook,
            skip_sheets={INPUT_DATA_SHEET_NAME} if INPUT_DATA_SHEET_NAME in workbook.sheetnames else set(),
        )
        export_path = Path(path)
        workbook.save(export_path)
        _restore_input_data_artifacts(
            export_path,
            input_data_extension,
            input_data_cached_formula_values(project),
        )

    def _select_export_active_sheet(self, workbook: Workbook) -> None:
        if INPUT_DATA_SHEET_NAME in workbook.sheetnames:
            select_single_sheet(workbook, INPUT_DATA_SHEET_NAME)
        elif "Comp_Stat" in workbook.sheetnames:
            select_single_sheet(workbook, "Comp_Stat")

    def _build_comp_stat_sheet(
        self,
        sheet: Worksheet,
        rows: list[ExportRow],
        case_count: int,
        voltage_fills: dict[str, PatternFill],
    ) -> None:
        max_col = max(2 + case_count, 3)
        style_sheet = StyleSheet(sheet)
        prepare_sheet(sheet, max_col=max_col)
        sheet.freeze_panes = "C3"

        ranges, last_row = self._write_styled_rows(sheet, style_sheet, rows, max_col, voltage_fills)
        trim_sheet(sheet, max_row=max(last_row, 2), max_col=max_col)

        if "cb" in ranges:
            first_row, last_range_row = ranges["cb"]
            add_cb_conditional_formatting(sheet, first_row=first_row, last_row=last_range_row, first_col=3, last_col=max_col)
        if "constants" in ranges:
            first_row, last_range_row = ranges["constants"]
            add_constants_conditional_formatting(sheet, first_row=first_row, last_row=last_range_row, first_col=3, last_col=max_col)
        add_duplicate_case_name_formatting(sheet, row=2, first_col=3, last_col=max_col)
        set_common_widths(sheet, max_col=max_col, first_width=18, second_width=14)

    def _build_layer_stat_sheet(self, sheet: Worksheet, rows: list[ExportRow], case_count: int) -> None:
        max_col = max(2 + case_count, 3)
        style_sheet = StyleSheet(sheet)
        prepare_sheet(sheet, max_col=max_col)
        sheet.freeze_panes = "C3"

        ranges, last_row = self._write_styled_rows(sheet, style_sheet, rows, max_col)
        trim_sheet(sheet, max_row=max(last_row, 2), max_col=max_col)

        if "layer" in ranges:
            first_row, last_range_row = ranges["layer"]
            add_layer_conditional_formatting(
                sheet,
                first_row=first_row,
                last_row=last_range_row,
                first_col=3,
                last_col=max_col,
                fill_scheme="layer",
            )
        if "sweep" in ranges:
            first_row, last_range_row = ranges["sweep"]
            add_layer_conditional_formatting(
                sheet,
                first_row=first_row,
                last_row=last_range_row,
                first_col=3,
                last_col=max_col,
                fill_scheme="sweep",
            )
        add_duplicate_case_name_formatting(sheet, row=2, first_col=3, last_col=max_col)
        set_common_widths(sheet, max_col=max_col, first_width=15, second_width=22)

    def _build_res_flux_sheet(
        self,
        sheet: Worksheet,
        rows: list[ExportRow],
        case_count: int,
        case_header_source: Worksheet,
    ) -> None:
        max_col = max(2 + case_count, 3)
        style_sheet = StyleSheet(sheet, case_header_source=case_header_source)
        prepare_sheet(sheet, max_col=max_col)
        sheet.freeze_panes = "C3"

        _ranges, last_row = self._write_styled_rows(sheet, style_sheet, rows, max_col)
        trim_sheet(sheet, max_row=max(last_row, 2), max_col=max_col)
        add_duplicate_case_name_formatting(sheet, row=2, first_col=3, last_col=max_col)
        set_common_widths(sheet, max_col=max_col, first_width=12, second_width=16)

    def _build_mm_blocks_sheet(
        self,
        sheet: Worksheet,
        rows: list[ExportRow],
        voltage_fills: dict[str, PatternFill],
    ) -> None:
        max_col = 8
        style_sheet = StyleSheet(sheet)
        prepare_sheet(sheet, max_col=max_col)

        _ranges, last_row = self._write_styled_rows(sheet, style_sheet, rows, max_col, voltage_fills)
        trim_sheet(sheet, max_row=max(last_row, 1), max_col=max_col)
        set_common_widths(sheet, max_col=max_col, first_width=22, second_width=12)
        for column in range(3, max_col + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 14

    def _write_styled_rows(
        self,
        sheet: Worksheet,
        style_sheet: StyleSheet,
        rows: list[ExportRow],
        max_col: int,
        voltage_fills: dict[str, PatternFill] | None = None,
    ) -> tuple[dict[str, tuple[int, int]], int]:
        ranges: dict[str, tuple[int, int]] = {}
        for row_index, row in enumerate(rows, start=1):
            if row.style_kind == "section":
                write_section_row(sheet, style_sheet, row_index, row.values, max_col, clear_fill=row.clear_fill)
            elif row.style_kind == "header_main":
                write_header_row(sheet, style_sheet, row_index, row.values, max_col, header_kind="main")
            elif row.style_kind == "header_mm":
                write_header_row(sheet, style_sheet, row_index, row.values, max_col, header_kind="mm")
            else:
                key_fill = None
                if voltage_fills is not None and row.style_kind in {"cb", "mm"}:
                    key_fill = voltage_fills.get(row.voltage)
                write_body_row(
                    sheet,
                    style_sheet,
                    row_index,
                    row.values,
                    max_col,
                    style_kind=row.style_kind,
                    position=row.position,
                    key_fill=key_fill,
                    force_black_font=row.force_black_font,
                )

            if row.range_key:
                first_row, _last_row = ranges.get(row.range_key, (row_index, row_index))
                ranges[row.range_key] = (first_row, row_index)

        return ranges, len(rows)


def _extract_input_data_extension(template_path: Path) -> bytes | None:
    sheet_part = _worksheet_part_for_sheet(template_path, INPUT_DATA_SHEET_NAME)
    if sheet_part is None:
        return None
    with ZipFile(template_path) as workbook_zip:
        worksheet_xml = workbook_zip.read(sheet_part)
    match = EXT_LST_RE.search(worksheet_xml)
    if match is None:
        return None
    return _ensure_extension_namespaces(match.group(0))


def _restore_input_data_artifacts(
    workbook_path: Path,
    extension_xml: bytes | None,
    cached_values: dict[str, object],
) -> None:
    if extension_xml is None and not cached_values:
        return
    sheet_part = _worksheet_part_for_sheet(workbook_path, INPUT_DATA_SHEET_NAME)
    if sheet_part is None:
        return

    fd, temp_name = tempfile.mkstemp(suffix=".xlsx", dir=str(workbook_path.parent))
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        with ZipFile(workbook_path) as source_zip, ZipFile(temp_path, "w", ZIP_DEFLATED) as target_zip:
            for item in source_zip.infolist():
                data = source_zip.read(item.filename)
                if item.filename == sheet_part:
                    if cached_values:
                        data = _replace_formula_cache(data, cached_values)
                    if extension_xml is not None:
                        data = _replace_or_insert_extension(data, extension_xml)
                target_zip.writestr(item, data)
        os.replace(temp_path, workbook_path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def _replace_or_insert_extension(worksheet_xml: bytes, extension_xml: bytes) -> bytes:
    without_extension = EXT_LST_RE.sub(b"", worksheet_xml)
    return without_extension.replace(b"</worksheet>", extension_xml + b"</worksheet>")


def _replace_formula_cache(worksheet_xml: bytes, cached_values: dict[str, object]) -> bytes:
    ET.register_namespace("", WORKBOOK_NS)
    root = ET.fromstring(worksheet_xml)
    formula_tag = f"{{{WORKBOOK_NS}}}f"
    value_tag = f"{{{WORKBOOK_NS}}}v"
    for cell in root.iter(f"{{{WORKBOOK_NS}}}c"):
        cell_ref = cell.attrib.get("r")
        if cell_ref not in cached_values:
            continue
        formula = cell.find(formula_tag)
        if formula is None:
            continue
        cell.attrib.pop("t", None)
        value = cell.find(value_tag)
        if value is None:
            value = ET.Element(value_tag)
            formula_index = list(cell).index(formula)
            cell.insert(formula_index + 1, value)
        value.text = _xml_scalar(cached_values[cell_ref])
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _xml_scalar(value: object) -> str:
    if isinstance(value, float):
        return f"{value:.15g}"
    return str(value)


def _ensure_extension_namespaces(extension_xml: bytes) -> bytes:
    if b"xr:" in extension_xml and b"xmlns:xr=" not in extension_xml:
        return extension_xml.replace(b"<ext ", f'<ext xmlns:xr="{XR_NS}" '.encode("utf-8"), 1)
    return extension_xml


def _worksheet_part_for_sheet(workbook_path: Path, sheet_name: str) -> str | None:
    with ZipFile(workbook_path) as workbook_zip:
        workbook_root = ET.fromstring(workbook_zip.read("xl/workbook.xml"))
        relationship_id = _sheet_relationship_id(workbook_root, sheet_name)
        if relationship_id is None:
            return None

        relationships_root = ET.fromstring(workbook_zip.read("xl/_rels/workbook.xml.rels"))
        for relationship in relationships_root:
            if relationship.attrib.get("Id") != relationship_id:
                continue
            if relationship.attrib.get("Type") != WORKSHEET_REL_TYPE:
                return None
            target = relationship.attrib.get("Target", "")
            if target.startswith("/"):
                return target.lstrip("/")
            return f"xl/{target}"
    return None


def _sheet_relationship_id(workbook_root: ET.Element, sheet_name: str) -> str | None:
    sheets = workbook_root.find(f"{{{WORKBOOK_NS}}}sheets")
    if sheets is None:
        return None
    for sheet in sheets:
        if sheet.attrib.get("name") == sheet_name:
            return sheet.attrib.get(f"{{{OFFICE_REL_NS}}}id")
    return None
