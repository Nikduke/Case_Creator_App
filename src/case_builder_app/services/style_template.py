from __future__ import annotations

import colorsys
from copy import copy
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Color, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from case_builder_app.services.equipment import voltage_sort_key, voltage_token


class StyleSheet:
    def __init__(self, sheet: Worksheet, *, case_header_source: Worksheet | None = None) -> None:
        self.sheet = sheet
        self.case_header_styles = {
            col: _capture_style(case_header_source.cell(2, col))
            for col in range(1, 4)
        } if case_header_source is not None else {}
        section_row = 33 if sheet.title == "Comp_Stat" else 1
        self.styles: dict[tuple[str, str, int], dict[str, Any]] = {}
        self._capture("section", "middle", section_row)
        self._capture("header_main", "middle", 2)
        self._capture("header_mm", "middle", 1)
        self._capture("cb", "top", 7)
        self._capture("cb", "middle", 8)
        self._capture("cb", "bottom", 9)
        self._capture("cb", "single", 9)
        self._capture("fault", "top", 34)
        self._capture("fault", "middle", 35)
        self._capture("fault", "bottom", 37)
        self._capture("fault", "single", 34)
        self._capture("constant_main", "middle", 39)
        self._capture("constant_other", "middle", 40)
        self._capture("layer", "top", 3)
        self._capture("layer", "middle", 4)
        self._capture("layer", "bottom", 19)
        self._capture("layer", "single", 3)
        self._capture("flux", "top", 3)
        self._capture("flux", "middle", 3)
        self._capture("flux", "bottom", 3)
        self._capture("flux", "single", 3)
        self.styles[("flux_value", "middle", 3)] = _capture_style(sheet["A3"])
        self._capture("mm", "top", 11)
        self._capture("mm", "middle", 12)
        self._capture("mm", "bottom", 24)
        self._capture("mm", "single", 24)
        self.group_fills = [
            copy(sheet["A11"].fill),
            copy(sheet["A3"].fill),
            copy(sheet["A21"].fill),
            copy(sheet["A7"].fill),
            copy(sheet["A24"].fill),
        ]

    def group_fill(self, index: int) -> PatternFill:
        if index < len(self.group_fills):
            return copy(self.group_fills[index])
        return _generated_group_fill(index - len(self.group_fills))

    def _capture(self, kind: str, position: str, row: int) -> None:
        for col in range(1, 4):
            self.styles[(kind, position, col)] = _capture_style(self.sheet.cell(row, col))

    def apply(self, cell, kind: str, col: int, *, position: str = "middle") -> None:  # noqa: ANN001
        template_col = _template_col(col)
        if kind == "header_main" and col > 2 and self.case_header_styles:
            _apply_captured_style(self.case_header_styles[template_col], cell)
            return
        if kind == "flux" and col > 2:
            _apply_captured_style(self.styles[("flux_value", "middle", 3)], cell)
            return
        style = self.styles.get((kind, position, template_col)) or self.styles.get((kind, "middle", template_col))
        if style is None:
            style = self.styles[("section", "middle", template_col)]
        _apply_captured_style(style, cell)


def build_voltage_fill_map(sheet: Worksheet, names: list[str]) -> dict[str, PatternFill]:
    style_sheet = StyleSheet(sheet)
    voltage_tokens = sorted({voltage_token(name) for name in names if name.strip()}, key=voltage_sort_key)
    return {voltage: style_sheet.group_fill(index) for index, voltage in enumerate(voltage_tokens)}


def prepare_sheet(sheet: Worksheet, *, max_col: int) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 100
    sheet.sheet_view.zoomScaleNormal = 100
    sheet.conditional_formatting = ConditionalFormattingList()
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None
            cell.comment = None
            cell.hyperlink = None
    for col in range(1, max_col + 1):
        sheet.column_dimensions[get_column_letter(col)].hidden = False


def trim_sheet(sheet: Worksheet, *, max_row: int, max_col: int) -> None:
    if sheet.max_row > max_row:
        sheet.delete_rows(max_row + 1, sheet.max_row - max_row)
    if sheet.max_column > max_col:
        sheet.delete_cols(max_col + 1, sheet.max_column - max_col)


def set_common_widths(sheet: Worksheet, *, max_col: int, first_width: float, second_width: float) -> None:
    sheet.column_dimensions["A"].width = first_width
    sheet.column_dimensions["B"].width = second_width
    for col in range(3, max_col + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 23
    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 17.25


def set_workbook_zoom(workbook: Workbook, *, skip_sheets: set[str] | None = None) -> None:
    skip_sheets = skip_sheets or set()
    for sheet in workbook.worksheets:
        if sheet.title in skip_sheets:
            continue
        sheet.sheet_view.zoomScale = 100
        sheet.sheet_view.zoomScaleNormal = 100


def select_single_sheet(workbook: Workbook, sheet_name: str) -> None:
    active_index = workbook.sheetnames.index(sheet_name)
    workbook.active = active_index
    if workbook.views:
        workbook.views[0].activeTab = active_index
        workbook.views[0].firstSheet = active_index
    for sheet in workbook.worksheets:
        sheet.sheet_view.tabSelected = None
    workbook[sheet_name].sheet_view.tabSelected = True


def write_section_row(
    sheet: Worksheet,
    style_sheet: StyleSheet,
    row: int,
    values: list[object],
    max_col: int,
    *,
    clear_fill: bool = False,
) -> None:
    for col in range(1, max_col + 1):
        cell = sheet.cell(row, col)
        style_sheet.apply(cell, "section", col)
        if clear_fill:
            cell.fill = PatternFill(fill_type=None)
        cell.value = values[col - 1] if col <= len(values) else None


def write_header_row(
    sheet: Worksheet,
    style_sheet: StyleSheet,
    row: int,
    values: list[object],
    max_col: int,
    *,
    header_kind: str = "main",
) -> None:
    for col in range(1, max_col + 1):
        cell = sheet.cell(row, col)
        style_sheet.apply(cell, f"header_{header_kind}", col)
        cell.value = values[col - 1] if col <= len(values) else None


def write_body_row(
    sheet: Worksheet,
    style_sheet: StyleSheet,
    row: int,
    values: list[object],
    max_col: int,
    *,
    style_kind: str,
    position: str,
    key_fill: PatternFill | None = None,
    force_black_font: bool = False,
) -> None:
    for col in range(1, max_col + 1):
        cell = sheet.cell(row, col)
        style_sheet.apply(cell, style_kind, col, position=position)
        if key_fill is not None and col <= 2:
            cell.fill = copy(key_fill)
        if style_kind == "fault":
            cell.fill = PatternFill(fill_type=None)
        if force_black_font:
            cell.font = _black_font(cell.font)
        if style_kind in {"cb", "mm"}:
            _apply_group_border(cell, position)
        if style_kind == "flux":
            _apply_flux_border(cell, position, col, max_col)
        cell.value = values[col - 1] if col <= len(values) else None


def add_cb_conditional_formatting(
    sheet: Worksheet,
    *,
    first_row: int,
    last_row: int,
    first_col: int,
    last_col: int,
) -> None:
    if last_col < first_col or last_row < first_row:
        return
    ref = _range_ref(first_row, last_row, first_col, last_col)
    anchor = sheet.cell(first_row, first_col).coordinate
    sheet.conditional_formatting.add(ref, _contains_text_rule(anchor, "OFF", _cf_fill("A63A3A"), _white_font()))
    sheet.conditional_formatting.add(ref, _contains_text_rule(anchor, "SWITCH", _cf_fill("E46C0A"), _white_font()))
    sheet.conditional_formatting.add(ref, _contains_text_rule(anchor, "ON", _cf_fill("77933C"), _white_font()))


def add_constants_conditional_formatting(
    sheet: Worksheet,
    *,
    first_row: int,
    last_row: int,
    first_col: int,
    last_col: int,
) -> None:
    if last_col < first_col or last_row < first_row:
        return
    ref = _range_ref(first_row, last_row, first_col, last_col)
    anchor = sheet.cell(first_row, first_col).coordinate
    sheet.conditional_formatting.add(ref, _contains_text_rule(anchor, "ON", _cf_fill("FF0000")))
    sheet.conditional_formatting.add(ref, _contains_text_rule(anchor, "OFF", _cf_fill("C6EFCE"), Font(color=_argb("006100"))))
    sheet.conditional_formatting.add(ref, _contains_text_rule(anchor, "ENERG", _cf_fill("FFC7CE")))


def add_layer_conditional_formatting(
    sheet: Worksheet,
    *,
    first_row: int,
    last_row: int,
    first_col: int,
    last_col: int,
    fill_scheme: str,
) -> None:
    if last_col < first_col or last_row < first_row:
        return
    disable_fill = _cf_fill("FFC7CE")
    enable_fill = _cf_fill("C6EFCE")
    if fill_scheme == "sweep":
        disable_fill = _cf_fill(_theme_color(9, 0.3999450666829432))
        enable_fill = _cf_fill(_theme_color(5, 0.0))
    ref = _range_ref(first_row, last_row, first_col, last_col)
    anchor = sheet.cell(first_row, first_col).coordinate
    sheet.conditional_formatting.add(ref, _contains_text_rule(anchor, "Disable", disable_fill))
    sheet.conditional_formatting.add(ref, _contains_text_rule(anchor, "Enable", enable_fill))


def add_duplicate_case_name_formatting(sheet: Worksheet, *, row: int, first_col: int, last_col: int) -> None:
    if last_col <= first_col:
        return
    ref = _range_ref(row, row, first_col, last_col)
    dxf = DifferentialStyle(fill=_cf_fill("FFC7CE"), font=Font(color=_argb("9C0006")))
    sheet.conditional_formatting.add(ref, Rule(type="duplicateValues", dxf=dxf))


def _capture_style(cell) -> dict[str, Any]:  # noqa: ANN001
    return {
        "style": copy(cell._style) if cell.has_style else None,
        "number_format": cell.number_format,
        "alignment": copy(cell.alignment),
        "protection": copy(cell.protection),
    }


def _apply_captured_style(style: dict[str, Any], target) -> None:  # noqa: ANN001
    if style["style"] is not None:
        target._style = copy(style["style"])
    target.number_format = style["number_format"]
    target.alignment = copy(style["alignment"])
    target.protection = copy(style["protection"])


def _template_col(col: int) -> int:
    return min(col, 3) if col > 2 else col


def _range_ref(first_row: int, last_row: int, first_col: int, last_col: int) -> str:
    return f"{get_column_letter(first_col)}{first_row}:{get_column_letter(last_col)}{last_row}"


def _contains_text_rule(anchor: str, text: str, fill: PatternFill, font: Font | None = None) -> Rule:
    return Rule(
        type="containsText",
        operator="containsText",
        text=text,
        dxf=DifferentialStyle(fill=fill, font=font),
        formula=[f'NOT(ISERROR(SEARCH("{text}",{anchor})))'],
    )


def _fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=_argb(color))


def _generated_group_fill(index: int) -> PatternFill:
    hue = (0.08 + index * 0.61803398875) % 1.0
    red, green, blue = colorsys.hls_to_rgb(hue, 0.84, 0.35)
    return _fill(f"{round(red * 255):02X}{round(green * 255):02X}{round(blue * 255):02X}")


def _cf_fill(color: str | Color) -> PatternFill:
    return PatternFill(fill_type="solid", bgColor=_color(color))


def _color(color: str | Color) -> Color:
    if isinstance(color, Color):
        return copy(color)
    return Color(rgb=_argb(color))


def _theme_color(theme: int, tint: float) -> Color:
    return Color(theme=theme, tint=tint)


def _white_font() -> Font:
    return Font(color=Color(theme=0))


def _black_font(font: Font) -> Font:
    updated = copy(font)
    updated.color = _argb("000000")
    return updated


def _argb(color: str) -> str:
    stripped = color.strip().lstrip("#").upper()
    if len(stripped) == 6:
        return f"FF{stripped}"
    return stripped


def _apply_group_border(cell, position: str) -> None:  # noqa: ANN001
    border = copy(cell.border)
    medium = Side(style="medium", color="000000")
    if position in {"top", "single"}:
        border.top = medium
    if position in {"bottom", "single"}:
        border.bottom = medium
    cell.border = border


def _apply_flux_border(cell, position: str, col: int, max_col: int) -> None:  # noqa: ANN001
    border = copy(cell.border)
    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")

    border.top = medium if position in {"top", "single"} else thin
    border.bottom = medium if position in {"bottom", "single"} else thin

    if col == 1:
        border.left = medium
        border.right = thin
    elif col == 2:
        border.left = thin
        border.right = None
    elif (col - 3) % 2 == 0:
        border.left = medium
        border.right = medium if col == max_col else thin
    else:
        border.left = thin
        border.right = medium

    cell.border = border
